from users.models import UserProfile, Location
from snaps.models import Branch
from django.contrib.auth.models import User
from snaps.helpers import create_bitsian
from django.db.models import Q
import openpyxl


def export_bitsians():
    branches = Branch.objects.all()
    c = 0
    wb = openpyxl.Workbook()
    f = open("export.txt", "w")
    for i in branches:
        try:
            f.write(f"{i.branch_code} - {i.branch_name}\n\n\n")
            ws = wb.create_sheet(title=f"{i.branch_code}", index=c)
            ws = wb.worksheets[c]
            c += 1
            colms = Location.objects.filter(branch=i).distinct("row").count()
            ascii = 65
            for j in range(0, colms):
                if j == 0:
                    ws[f"{chr(ascii)}1"] = "Sitting Row"
                else:
                    ws[f"{chr(ascii)}1"] = f"Standing Row {j}"
                ascii += 1
            ascii = 65
            for l in range(0, colms):
                if l == 0:
                    f.write(f"Sitting Row: ")
                else:
                    f.write(f"Standing Row {j}: ")
                locs = Location.objects.filter(Q(branch=i) & Q(row=l)).order_by("x")
                row = 2
                for k in locs:
                    if k.tag:
                        ws[f"{chr(ascii)}{row}"] = f"{k.tag.name}".title()
                        f.write(f"{k.tag.name}, ".title())
                    else:
                        ws[f"{chr(ascii)}{row}"] = "-"
                        f.write(f"-, ".title())
                    row += 1
                ascii += 1
                f.write("\n")
            f.write("\n\n\n")
        except Exception as e:
            print(e)

    f.close()
    wb.save("Exported_excel.xlsx")
    wb.close()

    f = open("list.txt", "w")
    for branch in Branch.objects.filter():
        f.write(f"{branch.branch_code,branch.students.filter().count()}")
    f.close()


def export_excel():
    branches = Branch.objects.all()
    c = 0
    row = 2
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws["A1"] = "Name"
    ws["B1"] = "BITS_ID"
    ws["C1"] = "Branch"
    for i in branches:
        try:
            locs = Location.objects.filter(Q(branch=i)).order_by("row")
            for l in locs:
                ws["A{}".format(row)] = l.tag.name
                ws["B{}".format(row)] = l.tag.bits_id
                ws["C{}".format(row)] = i.branch_code
                row += 1
        except Exception as e:
            print(e)

    wb.save("Exported_excel.xlsx")
    wb.close()


def run():
    export_excel()
