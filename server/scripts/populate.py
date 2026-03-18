from users.models import UserProfile
from snaps.models import Branch
from django.contrib.auth.models import User
from snaps.helpers import create_bitsian
from openpyxl import load_workbook

branch_full_names = {
  "A1": "B.E. Chemical Engineering",
  "A2": "B.E. Civil Engineering",
  "A3": "B.E. Electrical & Electronics Engineering",
  "A4": "B.E. Mechanical Engineering",
  "A5": "B. Pharmacy",
  "A7": "B.E.Computer Science",
  "A8": "B.E. Electronics & Instrumentation Engineering",
  "AA": "B.E. Electronics & Communication Engineering",
  "AB": "B.E. Manufacturing Engineering",
  "BXA1": "Dual B.E. Chemical Engineering",
  "BXA2": "Dual B.E. Civil Engineering",
  "BXA3": "Dual B.E. Electrical & Electronics Engineering",
  "BXA4": "Dual B.E. Mechanical Engineering",
  "BXA7": "Dual B.E. Computer Science",
  "BXA8": "Dual B.E. Electronics & Instrumentation Engineering",
  "BXAA": "Dual B.E. Electronics & Communication Engineering",
  "BXAB": "Dual B.E. Manufacturing Engineering",
  "B1": "MSc. Biological Sciences",
  "B2": "MSc. Chemistry",
  "B3": "MSc. Economics",
  "B4": "MSc. Mathematics",
  "B5": "MSc. Physics",
  "RMIT": "RMIT Collaboration Programme",
  "H106": "M.E (Mechanical Engineering)",
  "H141": "M.E (Design Engineering)",
  "H112": "M.E. (Software Systems)",
  "H103": "M.E. (Computer Science)",
  "H130": "M.E. Civil (Transportation Engineering)",
  "H143": "M.E. Civil (Structural Engineering)",
  "H144": "M.E. Civil (Infrastructure Systems)",
  "H155": "M.E. Environmental",
  "H101": "ME Chemical",
  "H129": "ME Biotechnology",
  "H124": "M.E. Communication Engineering",
  "H123": "M.E. (Microelectronics)",
  "H142": "M.E. (Manufacturing Systems Engineering)",
  "H140": "M.E. (Embedded Systems)",
  "H146": "M. Pharmacy (Pharmaceutics)",
  "H147": "M. Pharmacy (Pharmaceutical Chemistry)",
  "H153": "M. Pharmacy (Pharmacology)",
  "H154": "MBA",
  "D2": "M.Sc. Tech (General Studies)",
  "PHXH": "Ph.D Humanities & Social Science",
  "PHPH": "PhD. Pharmacy",
  "PHCS": "Ph.D Computer Science & Information Systems",
  "PHMG": "Ph.D Management",
  "PHEF": "Ph.D Economics & Finance",
  "PHCH": "Ph.D Chemistry",
  "H131": "M.E. (Power Electronics and Drives)",
  "PHEE": "Ph.D Electrical and Electronics Engineering",
  "PHPY": "Ph.D Physics",
  "PHMA": "Ph.D Mathematics",
  "PHBI": "Ph.D Biological Sciences",
  "PHME": "Ph.D Mechanical Engineering",
  "PHCE": "Ph.D Civil Engineering",
  "A7UB": "B.E. Computer Science (2+2 UB)",
  "AAIS": "B.E. Electronics and Communication Engineering (2+2 ISU)",
  "A3UB": "B.E. Electrical and Electronics Engineering (2+2 UB)",
  "A4RM": "B.E. Mechanical (RMIT)",
  "A2RM": "B.E. Civil Engineering (RMIT)",
  "PHXP": "Ph.D XP",
  "PHOF": "Ph.D OF",
  "PHDF": "Ph.D DF",
  "PHDP": "Ph.D DP",
  "PHXF": "Ph.D XF",
  "PHRF": "Ph.D RF",
  "PHRP": "Ph.D RP",
}

def populate_bitsians():
    wb = load_workbook("scripts/mess_list.xlsx")
    ws = wb["Sheet1"]
    for i in range(2, ws.max_row + 1):
        bits_id = ws.cell(row=i, column=2).value
        user_type = bits_id[4]
        prefix = ""
        branch_name = ""
        if user_type == "P":
            prefix = "p"
            branch_name = bits_id[4:8]
        elif user_type == "H":
            prefix = "h"
            branch_name = bits_id[4:8]
            if bits_id[:4] != "2025":
                continue
        else:
            prefix = "f"
            if bits_id[:4] == "2023":
                branch_name = bits_id[4:6]
            elif bits_id[:4] == "2022" and (bits_id[4] == "B"):
                if bits_id[6] == "P" or bits_id[6] == "T":
                    continue
                branch_name = f"BX{bits_id[6:8]}"
            elif bits_id[:4] == "2024" and bits_id[6:8] != "PS": # 2024A4CP0327P
              branch_name = bits_id[4:8]
            else:
                continue
        username = ws.cell(row=i, column=3).value
        email = f"{prefix}{bits_id[:4]}{bits_id[len(bits_id)-4:len(bits_id)]}@pilani.bits-pilani.ac.in"
        try:
            user = create_bitsian(
                username=f"{prefix}{bits_id[:4]}{bits_id[len(bits_id)-4:len(bits_id)]}",
                email=email,
                bits_id=bits_id,
                name=username,
                branch_name=branch_full_names[branch_name],
                branch_code=branch_name,
            )
            print(f"{i} : {user}")
        except Exception as e:
            print(f"{i} : {e}")


def run():
    populate_bitsians()
